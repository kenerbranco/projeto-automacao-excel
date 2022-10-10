"""
Ae3 - Motor/Axial (envelope)
Av+ - Motor/Axial
He3 - Motor/Radial (envelope)
Hv+ - Motor/Radial
Radial - Carcaça/Radial
Axial - Carcaça/Axial
LAe3+ - Lewa/Atuem/Axial(envelope)
LAv+ - Lewa/Atuem/Axial
LRe3+ - Lewa/Atuem/Radial(envelope)
LRv+ - Lewa/Atuem/Radial
"""
import openpyxl
maq = dict()
maq_final = list()

# Leitura arquivo .xlsx
book = openpyxl.load_workbook("Relatório Aptitude.xlsx")

# Atribuindo sheets em variáveis
index = book['index']
book.create_sheet('Final')

# maximo de linhas e colunas
max_lin = index.max_row
max_col = index.max_column

def extDados(l):
    lst = list()
    for la in range(l, l+5):
        cel = index.cell(row=la, column=7)
        val = cel.value
        lst.append(val)
    return lst


# Programa principal
for l in range(1, max_lin+1):
    for c in range(1, max_col):
        maq.clear()
        celula = index.cell(row=l, column=c)
        valor = str(celula.value).split()
        if 'Ae3' in valor:
            maq['Motor/Axial(envelope)'] = extDados(l)
            maq_final.insert(0, maq.copy())
        elif 'Av+' in valor:
            maq['Motor/Axial'] = extDados(l)
            maq_final.insert(1, maq.copy())
        elif 'He3' in valor:
            maq['Motor/Radial(envelope)'] = extDados(l)
            maq_final.insert(2, maq.copy())
        elif 'Hv+' in valor:
            maq['Motor/Radial'] = extDados(l)
            maq_final.insert(3, maq.copy())
        elif 'Radial' in valor:
            maq['Carcaça/Radial'] = extDados(l)
            maq_final.insert(4, maq.copy())
        elif 'Axial' in valor:
            maq['Carcaça/Axial'] = extDados(l)
            maq_final.insert(5, maq.copy())
        elif 'LAe3+' in valor:
            maq['Lewa/Atuem/Axial(envelope)'] = extDados(l)
            maq_final.insert(6, maq.copy())
        elif 'LAv+' in valor:
            maq['Lewa/Atuem/Axial'] = extDados(l)
            maq_final.insert(7, maq.copy())
        elif 'LRe3+' in valor:
            maq['Lewa/Atuem/Radial(envelope)'] = extDados(l)
            maq_final.insert(8, maq.copy())
        elif 'LRv+' in valor:
            maq['Leva/Atuem/Radial'] = extDados(l)
            maq_final.insert(9, maq.copy())

#for v in maq_final:
#    print(f'{v}')

final = book['Final']

for ponto in maq_final:
    for k, v in ponto.items():
        final.append([k])
        for medida in v:
            final.append([medida])

book.save('teste-saida.xlsx')
